import customtkinter
import CTkMessagebox
from customtkinter import CTkInputDialog
import threading          # Threading wird importiert, damit mehrere Prozesse gleichzeitig stattfinden können, damit es ein LIVE Bild der Kamera gibt dazu aber auch die Erkennung und kein Prozess den anderen beeinträchtigt.
import cv2                # Open-CV wird importiert, damit Gesichter erkannt werden können. 
import face_recognition   # Face-recognition wird importiert, damit mithilfe von Open-CV Gesichter erkennt werden können. 
import numpy as np        # Numpy wird importiert, damit Open CV Gesichter speichern speichern kann. 
import os                           # OS wird importiert, damit die Bilder gespeichert werden können und Datein erstellt werden können. 
import glob                         # Glob wird importiert, damit alle Bilder später gelöscht werden können die die Endung jpg haben.
import json                         # Json wird importiert, damit Dinge wie die Zeit und die Benutzerdaten gespeichert werden können und wieder aufrufbar sind. 
from datetime import datetime       # Datetime wird import, damit klar ist zu welchem Zeitpunkt man eincheckt und auscheckt. 
import time as t                    # Time wird importiert, da durch die time.sleep() Funktion der Counter realisiert wird. 
import pandas as pd                 # Pandas wird verwendet, damit die  Daten in Exel als Dataframe angelegt werden, also in einer Tabelle sortiert werden.
import openpyxl                     # Openpyxl wird verwendet damit Exel Datein erstellt und bearbeitet werden können. 
import tkinter as tk
from tkinter import ttk
import cv2
class Zeiterfassung:
    ADMIN_CODE = "PK"
    def __init__(self, root):

        customtkinter.set_appearance_mode("System")
        customtkinter.set_default_color_theme("blue")
        self.root = root
        self.root.title("Zeiterfassung")
        self.root.geometry("800x800")
        self.start_button = customtkinter.CTkButton(master=self.root, text="Einchecken", command=self.starte_gesichtserkennung)
        self.start_button.grid(row=0, column=0, pady=100, padx=80)
        self.neuesGesicht_button = customtkinter.CTkButton(master=self.root, text="Neuer Nutzer", command=self.neues_gesicht)
        self.neuesGesicht_button.grid(row=0, column=1, pady=100, padx=80)
        self.second_visit = customtkinter.CTkButton(master=self.root, text="Auschecken", command=self.auschecken)
        self.second_visit.grid(row=0, column=2, pady=100, padx=80)
        self.canvas = customtkinter.CTkCanvas(master=self.root, width=480, height=360, bg='#1D2529')
        self.canvas.create_text(240, 180, text="Kamera", fill='white', font=('Verdana', 20))
        self.canvas.grid(row=1, column=0, columnspan=4, pady=20)
        self.wipe_button = customtkinter.CTkButton(master=self.root, text="Daten entfernen", command=self.wipe_data)
        self.wipe_button.grid(row=1, column=4, pady=100, padx=80)
        self.manage_button = customtkinter.CTkButton(master=self.root, text="Verwalten", command=self.show_manage_interface)
        self.manage_button.grid(row=1, column=3, pady=10, padx=10)  # Adjust padding values

    def starte_gesichtserkennung(self):
        self.recognition_thread = threading.Thread(target=self.gesichtserkennung_thread)
        CTkMessagebox.CTkMessagebox(title="Info", message="Warte nun bitte 5 Sekunden vor der Kamera!")
        self.recognition_thread.start()

    def auschecken(self):
        self.second_visit_thread= threading.Thread(target=self.auschecken_thread)
        CTkMessagebox.CTkMessagebox(title="Info", message="Warte nun bitte 5 Sekunden vor der Kamera!")
        self.second_visit_thread.start()

    def gesichtserkennung_thread(self):
        person_names = np.load('data/person_names.npy')             # Daten für bereits vorhandene Gesichter werden geladen
        person_image_paths = np.load('data/person_image_paths.npy')

        with open('data/time.json', 'r') as file:   # time.json wird aufgerufen
            time = json.load(file)                  # Inhalt der Json wird in Variable "time" gespeichert
        
        gesichter_daten, bekannte_namen = self.load_known_faces(person_image_paths, person_names) # Der Inhalt der bereits vorhandenen Gesichtern wird der Variable gesichter_daten und die Namen in der Variable bekannte_namen gespeichert 
        video_capture = cv2.VideoCapture(0)         # Kamera-Port wird geöffnet 

        start_time = t.time()
        text=1
        height, width= 360, 480
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 2
            
        center_position = (width // 2, height // 2)     # Vorschaubild wird zentriert angezeigt
        text_size, _ = cv2.getTextSize(str(text), font, font_scale, 2)
        text_position = (center_position[0] - text_size[0] // 2, center_position[1] + text_size[1] // 2)
        color = (0, 0, 255)  
        while t.time() - start_time < 5:        # 5 Sekunden Timer wird eingebaut, bis die Kamera ein Bild macht
            
            ret, frame = video_capture.read()
            
            if not ret:
                break

            face_locations = face_recognition.face_locations(frame) # Gesicht wird im Bild gesucht

            for face_location in face_locations:
                top, right, bottom, left = face_location # Koordinaten des erkannten Gesichtes werden in face_location abgespeichert
                face_encoding = face_recognition.face_encodings(frame, [face_location])[0]  # Gesichtsdaten werden verschlüsselt mithilfe der Biblothek face_recognition und dann als face_encoding gespeichert

                Treffer = face_recognition.compare_faces(gesichter_daten, face_encoding, tolerance=0.6) # Die verschlüsselten Daten werden mit den bereits vorhandenen Gesichtern verglichen, den Variablen die eben die Daten aus der JSON zugeordnet bekommen haben
                name = "Nicht registriert"  

                if True in Treffer:
                    first_match_index = Treffer.index(True)    # Liste mit einem Index wird verwendet, da der Index die gleiche Rheinfolge in Treffer hat sowie in der Liste mit den bekannten Namen. Dadurch wird der Name dem Gesicht zugeordnet
                    name = bekannte_namen[first_match_index]
                    time[name] = datetime.now()                # Die akutlle Zeit "datenime.now" wird in der json Datei abgespeichert
                    with open('data/time.json', 'w') as file:
                        json.dump(time, file, default=lambda obj: obj.isoformat() if isinstance(obj, datetime) else None) # Das Datum muss ins ISO-Format umgwandelt werden, da json sonst das Datum nicht speichern konnte


            cv2.putText(frame, str(text), text_position, font, font_scale, color, 2, cv2.LINE_AA) # Counter wird auf dem Vorschaubild angezeigt
            text+=1
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)






            # photo = customtkinter.CTkImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())
            # photo = create_custom_tk_image(frame_rgb)
            photo = tk.PhotoImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())    # Das Bild wird zu einem PhotoImage verwandelt, wodurch es zu Wartezeiten kommt, tkInter das Bild jedoch zeigen kann
            self.canvas.create_image(0, 0, anchor=customtkinter.NW, image=photo)
            self.canvas.photo = photo  
            self.root.update()                                         # Das aktuelle Bild wird als Canvas in TKinter angezeigt
            # self.customtkinter.update()
            
            t.sleep(0.40)
            
            
        # self.customtkinter.update()
        self.root.update()
        self.canvas.delete("all")   # Vorschau wird zurückgesetzt, da die Funktion beendet ist 
        self.canvas.create_text(240, 180, text="Kamera", fill='white', font=('Verdana', 20)) # Kamera Beschrifftung für das Vorschaubild wird wieder eingefügt

        video_capture.release() # Die Kamera wird wieder freigegeben
        cv2.destroyAllWindows() # Alle noch von Open CV offenen Fenster werden geschlossen, damit es später nichtmehr zu Bugs kommen kann

    def auschecken_thread(self):
        person_names = np.load('data/person_names.npy')
        person_image_paths = np.load('data/person_image_paths.npy')

        with open('data/time.json', 'r') as file:
            time = json.load(file)

        gesichter_daten, bekannte_namen = self.load_known_faces(person_image_paths, person_names)
        video_capture = cv2.VideoCapture(0)
        
        start_time = t.time()
        text=1
        flag = 0

        self.root.update()
        self.canvas.delete("all")
        self.canvas.create_text(240, 180, text="Look at the camera", fill='white', font=('Verdana', 20))
        
        while t.time() - start_time < 5:
            ret, frame = video_capture.read()
            if not ret:
                break

            face_locations = face_recognition.face_locations(frame)

            for face_location in face_locations:
                top, right, bottom, left = face_location
                face_encoding = face_recognition.face_encodings(frame, [face_location])[0]

                Treffer = face_recognition.compare_faces(gesichter_daten, face_encoding, tolerance=0.6)
                name = "Unknown"

                if True in Treffer:
                    first_match_index = Treffer.index(True)
                    name = bekannte_namen[first_match_index]
                    flag=1

                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

                # Convert OpenCV BGR image to RGB format
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # Convert RGB image to PhotoImage format
                # photo = customtkinter.CTkImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())
                # photo =  create_custom_tk_image(frame_rgb)
                photo = tk.PhotoImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())
                # Set the image on the Canvas
                self.canvas.create_image(0, 0, anchor=customtkinter.NW, image=photo)
                self.canvas.photo = photo
                self.root.update()
                
                if name in time:
                    
                    old= time[name]
                    old_datetime = datetime.fromisoformat(old)
                    
                    current_datetime = datetime.now()
                    old_date = old_datetime.date()
                    current_date = current_datetime.date()

                    old_cur_date_string = old_date.strftime("%Y-%m-%d")
                    cur_date_string = current_date.strftime("%Y-%m-%d")
                    
                    time_difference = current_datetime - old_datetime

                    total_minutes = time_difference.total_seconds() / 60
                    total_hours = total_minutes / 60
                    minutes_str = str(int(total_minutes))
                    hours_str = str(int(total_hours))
                    
                    
                    
                    time_str = hours_str+" hours"+" "+minutes_str+" minutes"
                    data = [name, time_str, cur_date_string, old_cur_date_string]
                    print(data)
                    with open('data/time.json', 'r') as file:
                        time = json.load(file)

                    del time[name]

                    with open('data/time.json', 'w') as file:
                        json.dump(time, file, default=lambda obj: obj.isoformat() if isinstance(obj, datetime) else None)

                    # messagebox.showinfo("Total Visit Time",time_str, parent=newWin)
                    CTkMessagebox.CTkMessagebox(title="Info", message="Total Visit"+time_str)
                    


                    video_capture.release()
                    file_path = "data/report.xlsx"
                    try:
                        existing_df = pd.read_excel(file_path)
                        new_df = pd.concat([existing_df, pd.DataFrame([data], columns=["Name", "Time","Date of First visit", "Date of Second visit"])], ignore_index=True)
                        new_df.to_excel(file_path, index=False)
                        
                    except FileNotFoundError:
                        df = pd.DataFrame([data], columns=["Name", "Time","Date of First visit", "Date of Second visit"])
                        df.to_excel(file_path, index=False)
                    

                    self.root.update()
                    self.canvas.delete("all")
                    self.canvas.create_text(240, 180, text="Webcam", fill='white', font=('Verdana', 20))

                    break

                else:

                    # messagebox.showinfo("Failed", "Please visit first", parent=newWin)
                    CTkMessagebox.CTkMessagebox(title="Failed", message="Please visit first")
                    
                    


        

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        self.root.update()
        self.canvas.delete("all")
        self.canvas.create_text(240, 180, text="Webcam", fill='white', font=('Verdana', 20))

        if flag==0:
             CTkMessagebox.CTkMessagebox(title="Error", message="Kein Gesicht wurde erkannt!!!", icon="cancel")
        cv2.destroyAllWindows()
    def neues_gesicht(self):
        capture_thread = threading.Thread(target=self.neues_gesicht_thread)
        capture_thread.start()

    def neues_gesicht_thread(self):
        person_names = np.load('data/person_names.npy')
        person_image_paths = np.load('data/person_image_paths.npy')
        text=1
        height, width= 360, 480
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 2
            
        center_position = (width // 2, height // 2)
        text_size, _ = cv2.getTextSize(str(text), font, font_scale, 2)
        text_position = (center_position[0] - text_size[0] // 2, center_position[1] + text_size[1] // 2)
        color = (0, 0, 255)  
        


        video_capture = cv2.VideoCapture(0)
        frame = []
        start_time = t.time()
        
        while t.time() - start_time < 5:

            ret, frame = video_capture.read()
            copy_frame = frame.copy()
            # Convert OpenCV BGR image to RGB format
            cv2.putText(copy_frame, str(text), text_position, font, font_scale, color, 2, cv2.LINE_AA)
            frame_rgb = cv2.cvtColor(copy_frame, cv2.COLOR_BGR2RGB)
            # Convert RGB image to PhotoImage format
            # photo = customtkinter.CTkImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())
            # photo = create_custom_tk_image(frame_rgb)
            photo = tk.PhotoImage(data=cv2.imencode('.png', frame_rgb)[1].tobytes())
            # Set the image on the Canvas
            self.canvas.create_image(0, 0, anchor=customtkinter.NW, image=photo)
            self.canvas.photo = photo
            self.root.update()
            text+=1
            t.sleep(1)
                

            key = cv2.waitKey(1) & 0xFF
            if key == ord('s') or key == ord('S'):
                    break

            elif key == 27:
                break  # Check for the 'ESC' key to exit the loop



        inputwin = customtkinter.CTk()







        namedialogue = CTkInputDialog(text="Enter the person's name:", title="Capture New Face")
        name  = namedialogue.get_input()

        # name = simpledialog.askstring("Capture New Face", "Enter the person's name:", parent=newWin)
        with open('data/credentials.json', 'r') as file:
            cred = json.load(file)


        usernamedialogue = CTkInputDialog(text="Enter your username:", title="Login")
        username  = usernamedialogue.get_input()
        passworddialogue = CTkInputDialog(text="Enter your password:", title="Login")
        password  = passworddialogue.get_input()

    


        if username and password:
            cred[username] = password
            with open('data/credentials.json', 'w') as file:
                json.dump(cred, file)
        
        if name:
            filename = f"{name}_captured_image.jpg"
            filepath = self.save_image(frame, "images", filename)

            person_image_paths = np.append(person_image_paths, filepath)
            person_names = np.append(person_names, name)
            np.save('data/person_image_paths.npy', person_image_paths)
            np.save('data/person_names.npy', person_names)

                    
            # messagebox.showinfo("Success", "Name has been added sucessfully", parent=newWin)
            CTkMessagebox.CTkMessagebox(title="Success", message="Name has been added sucessfully")


        # newWin.destroy()
        inputwin.destroy()

        
        self.canvas.delete("all")
        self.canvas.create_text(240, 180, text="Webcam", fill='white', font=('Verdana', 20))        

        video_capture.release()
        cv2.destroyAllWindows()

    def wipe_data(self):
        self.wipe_data_thread = threading.Thread(target=self.wipe_data_Thread)
        self.wipe_data_thread.start()

    def wipe_data_Thread(self):
        removing_files = glob.glob('images/*.jpg')
        with open('data/time.json', 'r') as file:
            time = json.load(file)
        for i in removing_files:
            os.remove(i)

        person_names = np.load('data/person_names.npy')
        person_image_paths = np.load('data/person_image_paths.npy')

        person_names = np.resize(person_names, 0)
        person_image_paths = np.resize(person_image_paths, 0)

        np.save('data/person_image_paths.npy', person_image_paths)
        np.save('data/person_names.npy', person_names)

        time = {}

        with open('data/time.json', 'w') as file:
            json.dump(time, file, default=lambda obj: obj.isoformat() if isinstance(obj, datetime) else None)

        CTkMessagebox.CTkMessagebox(title="Info", message="Zurücksetzen war erfolgreich")
        # messagebox.showinfo("Info", "Zurücksetzen war erfolgreich", parent=newWin)

        file_path = "data/report.xlsx"
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

        # Clear data starting from the second row
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None

            wb.save(file_path)
            
        except Exception as e:
            print(f"Error: {e}")

        

    def save_image(self, image, directory, filename):
        if not os.path.exists(directory):
            os.makedirs(directory)

        filepath = os.path.join(directory, filename)
        cv2.imwrite(filepath, image)
        return filepath

    def load_known_faces(self, image_paths, names):
        gesichter_daten = []
        bekannte_namen = []

        for image_path, name in zip(image_paths, names):
            image = face_recognition.load_image_file(image_path)
            face_encodings = face_recognition.face_encodings(image)
            gesichter_daten.extend(face_encodings)
            bekannte_namen.extend([name] * len(face_encodings))

        return gesichter_daten, bekannte_namen
    

    def show_manage_interface(self):
        # Destroy existing widgets on the main interface
        self.start_button.destroy()
        self.neuesGesicht_button.destroy()
        self.wipe_button.destroy()
        self.second_visit.destroy()
        self.manage_button.destroy()
        self.canvas.destroy()

        # Create widgets for the manage interface
# self.second_visit = customtkinter.CTkButton(master=self.root, text="Auschecken", command=self.auschecken)
        # self.manage_label = tk.Label(self.root, text="Management Interface", font=('Helvetica', 16, 'bold'), foreground='#FFFFFF', background='#001F3F')
        self.manage_label = customtkinter.CTkLabel(self.root, text="Management Interface", font=('Helvetica', 16, 'bold'))
        

        # self.manage_label = ttk.Label(self.root, text="Management Interface", font=('Helvetica', 16, 'bold'), foreground='#FFFFFF', background='#001F3F')
        self.manage_label.grid(row=0, column=0, columnspan=3, pady=20)


        

                # Create a button to go back to the main interfaces
        self.second_visit = customtkinter.CTkButton(master=self.root, text="Auschecken", command=self.auschecken)
        
        # self.back_button = ttk.Button(self.root, text="Back to Main", command=self.show_main_interface, style="DarkButton.TButton")
        self.back_button = customtkinter.CTkButton(self.root, text="Back to Main", command=self.show_main_interface)
        self.back_button.grid(row=1, column=0,pady=100, padx=80)
        # self.login_button = ttk.Button(self.root, text="Login", command=self.Login, style="DarkButton.TButton")
        self.login_button = customtkinter.CTkButton(self.root, text="Login", command=self.Login)
        self.login_button.grid(row=1, column=1, pady=100, padx=80)
        # self.reset_button = ttk.Button(self.root, text="Reset Password", command=self.reset_pass, style="DarkButton.TButton")
        self.reset_button = customtkinter.CTkButton(self.root, text="Reset Password", command=self.reset_pass)
        self.reset_button.grid(row=1, column=2, pady=100, padx=80)


    def show_main_interface(self):
        # Destroy existing widgets on the manage interface
        self.manage_label.destroy()
        self.back_button.destroy()
        self.login_button.destroy()
        self.reset_button.destroy()
        
        self.start_button = customtkinter.CTkButton(master=self.root, text="Einchecken", command=self.starte_gesichtserkennung)
        self.start_button.grid(row=0, column=0, pady=100, padx=80)

        self.neuesGesicht_button = customtkinter.CTkButton(master=self.root, text="Neuer Nutzer", command=self.neues_gesicht)
        self.neuesGesicht_button.grid(row=0, column=1, pady=100, padx=80)



        self.second_visit = customtkinter.CTkButton(master=self.root, text="Auschecken", command=self.auschecken)
        self.second_visit.grid(row=0, column=2, pady=100, padx=80)


        self.canvas = customtkinter.CTkCanvas(master=self.root, width=480, height=360, bg='#1D2529')
        self.canvas.create_text(240, 180, text="Kamera", fill='white', font=('Verdana', 20))
        self.canvas.grid(row=1, column=0, columnspan=4, pady=20)

        self.wipe_button = customtkinter.CTkButton(master=self.root, text="Daten entfernen", command=self.wipe_data)
        self.wipe_button.grid(row=1, column=4, pady=100, padx=80)

        self.manage_button = customtkinter.CTkButton(master=self.root, text="Verwalten", command=self.show_manage_interface)
        self.manage_button.grid(row=1, column=3, pady=10, padx=10)  # Adjust padding values









        

    
    def reset_pass(self):
        
        try:


            inputwin = customtkinter.CTk()
            namedialogue = CTkInputDialog(text="Enter your username:", title="Login")
            username = namedialogue.get_input()
            admincodedialogue = CTkInputDialog(text="Enter admin code:", title="Login")
            admincode = admincodedialogue.get_input()
            passdialogue = CTkInputDialog(text="Enter new password:", title="Login")
            password = passdialogue.get_input()



            # Validate admin code
            if admincode == self.ADMIN_CODE:
                # Validate if all inputs are provided
                if username and password:
                    # Load existing credentials from JSON file
                    with open('data/credentials.json', 'r') as file:
                        cred = json.load(file)

                    # Update credentials with the new password
                    cred[username] = password

                    # Save the updated credentials back to the file
                    with open('data/credentials.json', 'w') as file:
                        json.dump(cred, file)

                    # Display a success message
                    # messagebox.showinfo(title="Info", message="Das passwort wurde zurückgesetzt",parent=newWin)
                    CTkMessagebox.CTkMessagebox(title="Info", message="Das passwort wurde zurückgesetzt")
                else:
                    # Display an error message if any input is missing
                    # messagebox.showinfo(title="Info", message="Fehlende Infos",parent=newWin)
                    CTkMessagebox.ctkmessagebox(title="Info", message="Fehlende Infos")
            else:
                # Display an error message if the admin code is incorrect
                # messagebox.showinfo(title="Info", message="Admin-Code falsch",parent=newWin)
                CTkMessagebox.CTkMessagebox(title="Info", message="Admin-Code falsch")

        except Exception as e:
            # Handle any unexpected errors and display an error message
            # messagebox.showinfo(title="Info", message="Fehler",parent=newWin)
            CTkMessagebox.CTkMessagebox(title="Info", message="Fehler")
        inputwin.destroy()



        

    # ... (existing code)
    def Login(self):


        inputwin = customtkinter.CTk()
        namedialogue = CTkInputDialog(text="Enter your username:", title="Login")
        username = namedialogue.get_input()
        passdialogue = CTkInputDialog(text="Enter new password:", title="Login")
        password = passdialogue.get_input()
        

        
        
        with open('data/credentials.json', 'r') as file:
            cred = json.load(file)
        if cred.get(username) == password:  # Use cred.get(username) to handle cases where the username is not in the dictionary
            self.display_filtered_rows(username)
            inputwin.destroy()
        else:
            CTkMessagebox.CTkMessagebox(title="Error", message="Falsches Passwort oder Nutzername!!!", icon="cancel")
            # messagebox.showerror("Error", "Invalid username or password")
            inputwin.destroy()

    def display_filtered_rows(self, username):
    # Clear existing content on the canvas
        self.manage_label.destroy()
        self.back_button.destroy()
        self.login_button.destroy()
        self.reset_button.destroy()





        self.canvas = tk.Canvas(self.root, width=480, height=360, bg='#003366')
        self.canvas.grid(row=0, column=0, sticky="nsew")



        # Add a frame inside the canvas to hold the Excel data
        self.data_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.data_frame, anchor="nw")

        self.load_button = ttk.Button(self.root, text="Load Excel", command=lambda: self.load_excel(username))
        self.load_button.grid(row=1, column=0)

        self.logout_button = ttk.Button(self.root, text="Logout", command=self.show_main_interface2)
        self.logout_button.grid(row=2, column=0)

        # Set row and column weights to make the canvas and scrollbar expandable
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)





















        # self.canvas = tk.Canvas(self.root, width=480, height=360, bg='#003366')
        # self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # # Add a vertical scrollbar to the canvas
        # vscrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        # vscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        # self.canvas.configure(yscrollcommand=vscrollbar.set)

        # # Add a frame inside the canvas to hold the Excel data
        # self.data_frame = tk.Frame(self.canvas)
        # self.canvas.create_window((0, 0), window=self.data_frame, anchor="nw")
        
        # self.load_button = ttk.Button(self.root, text="Load Excel", command=self.load_excel(username))
        # # self.load_button = customtkinter.CTKButton(self.root, text="Load Excel", command=self.load_excel(username))
        # self.load_button.pack()
        
        # self.logout_button = ttk.Button(self.root, text="Logout", command=self.show_main_interface2)
        # # self.logout_button = customtkinter.CTKButton(self.root, text="Logout", command=self.show_main_interface2)
        # self.logout_button.pack()

        # vscrollbar.pack_forget()
        


    def load_excel(self, username):
        file_path = 'data/report.xlsx'

        if file_path:
            try:
                # Read Excel data using pandas
                df = pd.read_excel(file_path)

                # Display data on the canvas
                self.display_data(df, username)

            except Exception as e:
                CTkMessagebox.CTkMessagebox(title="Error", message="Error loading Excel file")
                # messagebox.showerror("Error", f"Error loading Excel file: {e}")

    def display_data(self, data_frame, target_name):
        # Clear previous content on the canvas
        self.data_frame.destroy()

        # Create a new frame to hold the Excel data
        self.data_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.data_frame, anchor="nw")

        # Convert dataframe to a table format (for simplicity, you can customize this)
        table = ttk.Treeview(self.data_frame, columns=list(data_frame.columns), show="headings", height=10)

        # Add columns to the treeview
        for col in data_frame.columns:
            table.heading(col, text=col)
            table.column(col, width=100, anchor="center")  # Adjust width as needed

        # Add rows to the treeview based on the 'Name' column
        for index, row in data_frame.iterrows():
            if target_name=='admin':
                table.insert("", "end", values=list(row))
            elif row['Name'] == target_name:
                table.insert("", "end", values=list(row))
                

        # Add the treeview to the frame
        table.pack(side="left", fill="both", expand=True)

        # Update the canvas scroll region
        self.canvas.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def show_main_interface2(self):

                # Destroy existing widgets on the manage interface
        self.manage_label.destroy()
        self.back_button.destroy()
        self.login_button.destroy()
        self.reset_button.destroy()
        self.load_button.grid_forget()
        self.logout_button.grid_forget()
        self.canvas.grid_forget()
        self.data_frame.grid_forget()
        # self.vscrollbar.grid_forget()
        
        self.start_button = customtkinter.CTkButton(master=self.root, text="Einchecken", command=self.starte_gesichtserkennung)
        self.start_button.grid(row=0, column=0, pady=100, padx=80)

        self.neuesGesicht_button = customtkinter.CTkButton(master=self.root, text="Neuer Nutzer", command=self.neues_gesicht)
        self.neuesGesicht_button.grid(row=0, column=1, pady=100, padx=80)



        self.second_visit = customtkinter.CTkButton(master=self.root, text="Auschecken", command=self.auschecken)
        self.second_visit.grid(row=0, column=2, pady=100, padx=80)


        self.canvas = customtkinter.CTkCanvas(master=self.root, width=480, height=360, bg='#1D2529')
        self.canvas.create_text(240, 180, text="Kamera", fill='white', font=('Verdana', 20))
        self.canvas.grid(row=1, column=0, columnspan=4, pady=20)

        self.wipe_button = customtkinter.CTkButton(master=self.root, text="Daten entfernen", command=self.wipe_data)
        self.wipe_button.grid(row=1, column=4, pady=100, padx=80)

        self.manage_button = customtkinter.CTkButton(master=self.root, text="Verwalten", command=self.show_manage_interface)
        self.manage_button.grid(row=1, column=3, pady=10, padx=10)  # Adjust padding values





if __name__ == "__main__":
    root = customtkinter.CTk()
    Zeiterfassung(root)
    root.mainloop()
